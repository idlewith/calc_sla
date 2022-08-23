"""
读excel文件
读前两列，获取开始、结束时间
是否周末，是，减去24小时
是否法定假日，是，减去24小时
是正常工作日，计算中午、下班时间是否在上班时间范围内，适当减去休息时间

示例：
创建日期	关闭日期	总解决时间（min）	总解决时间（h）	最终结果（h）
2022/5/16 08:28	2022/5/16 17:06	518	8.6	7.1
2022/5/16 16:55	2022/5/17 10:46	1071	17.9	2.9
2022/5/20 16:54	2022/5/23 17:15	4340	72.3	7.9
2022/5/3 09:30	2022/5/5 14:05	3155	52.6	4.1
"""
import datetime
import json
import os
from argparse import Namespace

import chinese_calendar as calendar
import openpyxl


def load_break_time():
    python_break_dict = {
        "launch_break_start": "11:30:00",
        "launch_break_end": "13:30:00",
        "launch_break_delta_hour": 1.5,
        "off_duty_first_start": "17:30:00",
        "off_duty_first_end": "23:59:59",
        "off_duty_second_start": "00:00:00",
        "off_duty_second_end": "08:30:00",
        "input_excel_file_name": "template.xlsx",
        "input_excel_sheet_name": "Sheet1",
        "output_excel_file_name": "result.xlsx",
    }
    json_file = "config.json"
    if not os.path.exists(json_file):
        used_break_dict = python_break_dict
    else:
        json_break_dict = json.load(open(json_file))
        same_dict = python_break_dict == json_break_dict
        if same_dict:
            used_break_dict = python_break_dict
        else:
            used_break_dict = json_break_dict
    used_break_dict = Namespace(**used_break_dict)
    return used_break_dict


BREAK_DICT = load_break_time()


class SlaModel:
    def __init__(self):
        # 创建日期: 2022/5/16 08:28
        self.start_time = ""

        # 关闭日期: 2022/5/16 17:06
        self.closed_time = ""

        # 总解决时间(min) 518
        self.total_time_delta_min = ""

        # 总解决时间(h) 8.6
        self.total_time_delta_hour = ""

        # 最终结果(h) 7.1
        self.actual_time_delta_hour = ""


def str2datetime(date_str):
    if isinstance(date_str, datetime.datetime):
        return date_str

    time_template = "%Y/%m/%d %H:%M:%S"
    a = datetime.datetime.strptime(date_str, time_template)
    return a


def calc_seconds(start_time, closed_time):
    delta = (closed_time - start_time).total_seconds()
    return delta


def get_hms(date):
    h, m, s = date.hour, date.minute, date.second
    return h, m, s


def gen_datetime_from_break_hms(date_day, break_time_str):
    y, m, d = get_ymd(date_day)

    time_template = "%H:%M:%S"
    break_start_hms = datetime.datetime.strptime(break_time_str, time_template)
    h, mi, s = get_hms(break_start_hms)
    return datetime.datetime(y, m, d, h, mi, s)


def is_in_launch_break_time(start, close):
    """
    only focus this: close time contains launch break end
    break time:     |----------------|
    actual:     |-------------------------|
    """

    launch_break_start = gen_datetime_from_break_hms(
        start, BREAK_DICT.launch_break_start
    )
    launch_break_end = gen_datetime_from_break_hms(start, BREAK_DICT.launch_break_end)
    if start <= launch_break_start and close >= launch_break_end:
        return True, BREAK_DICT.launch_break_delta_hour * 60 * 60
    elif launch_break_start <= start <= launch_break_end:
        return True, (launch_break_end - start).total_seconds()
    else:
        return False, 0


def is_in_off_duty_time(start, close):
    """
    在下班时间内解决了，就不能减去下班时间，
    只有整个事件包含了整个下班时间
        或者开始时间 > 下班时间并且结束时间 > 上班时间，才能减去下班时间

    break time:     |-------||---------|
    actual:     |-----------||------------|

    break time: |---------||-------|
    actual:         |-----||-------------------|
    """
    off_duty_first_start = gen_datetime_from_break_hms(
        start, BREAK_DICT.off_duty_first_start
    )
    off_duty_first_end = gen_datetime_from_break_hms(
        start, BREAK_DICT.off_duty_first_end
    )
    off_duty_second_start = gen_datetime_from_break_hms(
        close, BREAK_DICT.off_duty_second_start
    )
    off_duty_second_end = gen_datetime_from_break_hms(
        close, BREAK_DICT.off_duty_second_end
    )

    # 如果事件开始时间 < 下班时间
    if start <= off_duty_first_start and off_duty_first_end == close:
        delta_off_duty_first = (
            off_duty_first_end - off_duty_first_start
        ).total_seconds()
        # 如果事件开始事件 < 上班时间
        if off_duty_second_start <= start <= off_duty_second_end:
            delta_off_duty_early = (off_duty_second_end - start).total_seconds()
        else:
            delta_off_duty_early = 0
        delta_part = delta_off_duty_first + delta_off_duty_early
        return True, delta_part

    if start > off_duty_first_start and off_duty_first_end == close:
        return True, (off_duty_first_end - start).total_seconds()

    if start == off_duty_second_start and off_duty_second_end <= close:
        return True, (off_duty_second_end - off_duty_second_start).total_seconds()

    return False, 0


def calc_break_time(start, close):
    need_minus_seconds_list = []

    in_launch_break_time, delta_minus_seconds_launch = is_in_launch_break_time(
        start, close
    )
    if in_launch_break_time:
        need_minus_seconds_list.append(delta_minus_seconds_launch)

    in_off_duty_time, delta_minus_seconds_duty = is_in_off_duty_time(start, close)
    if in_off_duty_time:
        need_minus_seconds_list.append(delta_minus_seconds_duty)

    return sum(need_minus_seconds_list)


def calc_holiday_time(start, close):
    return (close - start).total_seconds()


def calc(start, close):
    ys, ms, ds = get_ymd(start)
    date_ymd = datetime.date(ys, ms, ds)
    holiday = calendar.is_holiday(date_ymd)
    workday = calendar.is_workday(date_ymd)

    if holiday or not workday:
        delta = calc_holiday_time(start, close)
    else:
        delta = calc_break_time(start, close)

    return delta


def is_same_day(start, close):
    s = str2datetime(start)
    c = str2datetime(close)
    return get_ymd(s) == get_ymd(c)


def get_ymd(date):
    y, m, d = date.year, date.month, date.day
    return y, m, d


def concat_start_day_list(date_start):
    y, m, d = get_ymd(date_start)
    end_of_day = datetime.datetime(y, m, d, 23, 59, 59)
    return [date_start, end_of_day]


def concat_close_day_list(date_end):
    y, m, d = get_ymd(date_end)
    start_of_day = datetime.datetime(y, m, d, 0, 0, 0)
    return [start_of_day, date_end]


def concat_whole_day_list(date_middle):
    y, m, d = get_ymd(date_middle)
    start_of_day = datetime.datetime(y, m, d, 0, 0, 0)
    end_of_day = datetime.datetime(y, m, d, 23, 59, 59)
    return [start_of_day, end_of_day]


def calc_days(s, c):
    days = (datetime.datetime(c.year, c.month, c.day) - datetime.datetime(
        s.year, s.month, s.day
    )).days
    return days


def get_date_list_from_more_than_one_day(start, close):
    date_list = []

    s = str2datetime(start)
    c = str2datetime(close)
    days = calc_days(s, c)

    start_day_list = concat_start_day_list(s)
    date_list.append(start_day_list)

    for i in range(1, days):
        day_middle = c - datetime.timedelta(days=i)
        whole_day_list = concat_whole_day_list(day_middle)
        date_list.append(whole_day_list)

    close_day_list = concat_close_day_list(c)
    date_list.append(close_day_list)
    date_list = sorted(date_list, key=lambda x: x[0])

    return date_list


def calc_single_event(start, close):
    sla_model = SlaModel()
    sla_model.start_time = start
    sla_model.closed_time = close

    summary_seconds = (str2datetime(close) - str2datetime(start)).total_seconds()

    same_day = is_same_day(start, close)
    if same_day:
        date_list = [[str2datetime(start), str2datetime(close)]]
    else:
        date_list = get_date_list_from_more_than_one_day(start, close)

    delta_list = []
    for start_close in date_list:
        start, close = start_close
        delta_time = calc(start, close)
        delta_list.append(delta_time)
    delta_summary_seconds = sum(delta_list)
    actual_seconds = summary_seconds - delta_summary_seconds

    sla_model.total_time_delta_min = summary_seconds / 60
    sla_model.total_time_delta_hour = sla_model.total_time_delta_min / 60
    sla_model.actual_time_delta_hour = actual_seconds / 3600
    return sla_model


def calc_multiple_events(date_list):
    # date_list = [
    #     ['2022/5/16 08:28:05', '2022/5/16 17:06:22'],
    #     ['2022/5/16 16:55:17', '2022/5/17 10:46:11'],
    #     ['2022/5/20 16:54:42', '2022/5/23 17:15:02'],
    #     ['2022/5/3 09:30:32', '2022/5/5 14:05:41'],
    # ]

    r = []
    for start_close in date_list:
        start, close = start_close
        sla_model: SlaModel = calc_single_event(start, close)
        r.append(
            [
                sla_model.start_time.strftime("%Y/%m/%d %H:%M:%S"),
                sla_model.closed_time.strftime("%Y/%m/%d %H:%M:%S"),
                f"{sla_model.total_time_delta_min:.0f}",
                f"{sla_model.total_time_delta_hour:.1f}",
                f"{sla_model.actual_time_delta_hour:.1f}",
            ]
        )
    return r


def read_excel(file_excel, sheet_name):
    wb = openpyxl.load_workbook(file_excel)
    sheet_names = wb.sheetnames
    if sheet_name not in sheet_names:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    result = []
    for row_index in range(max_row):
        row_index = row_index + 1

        row_list = []
        for col_index in range(max_col):
            col_index = col_index + 1

            cell_value = ws.cell(row=row_index, column=col_index).value
            row_list.append(cell_value)
        result.append(row_list)
    wb.close()
    return result


def write_to_excel(file_excel, data_list):
    data_list = list(data_list)

    wb = openpyxl.Workbook()
    ws = wb.active

    for row_index, row_list in enumerate(data_list):
        row_index = row_index + 1

        for col_index, cell_value in enumerate(row_list):
            col_index = col_index + 1

            ws.cell(row=row_index, column=col_index, value=cell_value)

    wb.save(file_excel)


def main():
    excel_list = read_excel(
        BREAK_DICT.input_excel_file_name, BREAK_DICT.input_excel_sheet_name
    )
    date_list = []
    for row in excel_list[1:]:
        start, close = row[0], row[1]
        if start and close:
            date_list.append([start, close])
    data_list = [excel_list[0]]
    result_list = calc_multiple_events(date_list)
    data_list.extend(result_list)
    write_to_excel(BREAK_DICT.output_excel_file_name, data_list)


if __name__ == "__main__":
    main()
